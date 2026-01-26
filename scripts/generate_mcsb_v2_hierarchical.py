#!/usr/bin/env python3
"""
Generador Excel MCSB v2 con estructura jerárquica padre-hijo
Extrae de https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-*
Incluye Azure Policy mapping de mcsb-v2-controls-policy-mapping
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
import time

# Cache global de Azure Policies (con URLs)
AZURE_POLICIES_CACHE = {}  # {control_id: [(name, url), ...]}

# Cache global de Core Pillars por dominio
CORE_PILLARS_CACHE = {}  # {domain_prefix: {control_id: pillar_name}}

# Dominios MCSB v2
DOMAINS = [
    ("Network Security", "network-security", "NS"),
    ("Identity Management", "identity-management", "IM"),
    ("Privileged Access", "privileged-access", "PA"),
    ("Data Protection", "data-protection", "DP"),
    ("Asset Management", "asset-management", "AM"),
    ("Logging and Threat Detection", "logging-threat-detection", "LT"),
    ("Incident Response", "incident-response", "IR"),
    ("Posture and Vulnerability Management", "posture-vulnerability-management", "PV"),
    ("Endpoint Security", "endpoint-security", "ES"),
    ("Backup and Recovery", "backup-recovery", "BR"),
    ("DevOps Security", "devops-security", "DS"),
    ("Artificial Intelligence Security", "artificial-intelligence-security", "AI")
]

# Mapeo de prefix → domain_slug para hyperlinks
DOMAIN_SLUG_MAP = {prefix: slug for _, slug, prefix in DOMAINS}

def load_azure_policies():
    """Carga el mapping de Azure Policies desde la página oficial"""
    global AZURE_POLICIES_CACHE

    if AZURE_POLICIES_CACHE:
        return  # Ya cargado

    print(f"\n{'='*80}")
    print("Cargando Azure Policy Mappings...")
    print(f"{'='*80}")

    url = "https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-controls-policy-mapping"

    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'lxml')

        # Buscar H2 (controles) y sus tablas
        h2_elements = soup.find_all('h2')

        for h2 in h2_elements:
            h2_text = h2.get_text(strip=True)

            # Detectar control ID (ej: "AI-1: Title")
            control_match = re.match(r'^([A-Z]{2}-\d+)[:\s]', h2_text, re.IGNORECASE)
            if not control_match:
                continue

            control_id = control_match.group(1).upper()

            # Buscar tabla asociada
            current = h2.find_next_sibling()
            while current and current.name != 'h2':
                if current.name == 'table':
                    rows = current.find_all('tr')
                    policies = []  # Lista de tuplas (name, url)

                    for row in rows[1:]:  # Skip header
                        cells = row.find_all('td')
                        if cells:
                            link = cells[0].find('a')
                            if link:
                                policy_name = link.get_text(strip=True)
                                policy_url = link.get('href', '')
                                if policy_name:
                                    policies.append((policy_name, policy_url))
                            else:
                                policy_name = cells[0].get_text(strip=True)
                                if policy_name:
                                    policies.append((policy_name, ''))

                    if policies:
                        AZURE_POLICIES_CACHE[control_id] = policies
                        print(f"  {control_id}: {len(policies)} policies")
                    break

                current = current.find_next_sibling()

        print(f"  Total controles con policies: {len(AZURE_POLICIES_CACHE)}")

    except Exception as e:
        print(f"  ⚠️  Error cargando policies: {str(e)}")

def extract_core_pillars(domain_slug, domain_prefix):
    """Extrae core pillars de la introducción de cada dominio"""
    global CORE_PILLARS_CACHE

    url = f"https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-{domain_slug}"

    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'lxml')

        pillar_mapping = {}  # {control_id: pillar_name}
        current_pillar = None

        # Buscar todos los elementos <p>
        for elem in soup.find_all('p'):
            text = elem.get_text(strip=True)

            # Detectar nombre de pillar: empieza con mayúscula, contiene ":", y es descriptivo
            # Ejemplos: "Secure network boundaries:", "Apply network isolation:", etc.
            if ':' in text and text[0].isupper() and not text.startswith('Note'):
                # Extraer la parte antes de los dos puntos
                pillar_candidate = text.split(':', 1)[0].strip()

                # Validar longitud y que no sea "Related controls"
                if 5 < len(pillar_candidate) < 60 and pillar_candidate != 'Related controls':
                    # Verificar que no sea una frase completa (evitar descripciones largas)
                    # Los pillar names típicamente tienen 2-5 palabras
                    word_count = len(pillar_candidate.split())
                    if 2 <= word_count <= 6:
                        current_pillar = pillar_candidate
                        continue

            # Detectar "Related controls:" seguido de lista
            if current_pillar and text == 'Related controls:':
                # El siguiente hermano debe ser <ul> con lista de controles
                next_ul = elem.find_next_sibling()
                if next_ul and next_ul.name == 'ul':
                    for li in next_ul.find_all('li'):
                        control_text = li.get_text(strip=True)
                        # Extraer control ID (ej: "NS-1: Title")
                        control_match = re.match(rf'^({domain_prefix}-\d+)[:\s]', control_text)
                        if control_match:
                            control_id = control_match.group(1).upper()
                            pillar_mapping[control_id] = current_pillar

                current_pillar = None  # Reset para siguiente pillar

        CORE_PILLARS_CACHE[domain_prefix] = pillar_mapping
        return pillar_mapping

    except Exception as e:
        print(f"    ⚠️  Error extrayendo core pillars: {str(e)}")
        return {}

def extract_controls_from_domain(domain_name, domain_slug, domain_prefix):
    """Extrae controles padre e hijos de un dominio MCSB v2"""
    url = f"https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-{domain_slug}"
    print(f"\n{'='*80}")
    print(f"Procesando: {domain_name} ({url})")
    print(f"{'='*80}")

    # Extraer core pillars primero
    pillar_mapping = extract_core_pillars(domain_slug, domain_prefix)
    if pillar_mapping:
        print(f"  ✓ Core pillars: {len(set(pillar_mapping.values()))} pillars mapeados a {len(pillar_mapping)} controles")

    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'lxml')

        controls = []
        current_parent = None

        # Buscar todos los H2 (controles padre)
        h2_elements = soup.find_all('h2')

        for h2 in h2_elements:
            h2_text = h2.get_text(strip=True)

            # Detectar controles padre: formato "NS-1: Title"
            control_match = re.match(rf'^({domain_prefix}-\d+)[:\s]+(.*)', h2_text, re.IGNORECASE)
            if not control_match:
                continue

            control_id = control_match.group(1).upper()
            control_title = control_match.group(2).strip()

            print(f"  [PADRE] {control_id}: {control_title}")

            # Obtener core pillar del control
            pillar_map = CORE_PILLARS_CACHE.get(domain_prefix, {})
            core_pillar = pillar_map.get(control_id, '')

            parent_data = {
                'Control ID': control_id,
                'Implementation ID': '',  # Vacío para padres
                'Control Name': control_title,
                'Control Type': 'Parent',
                'Core Pillar': core_pillar,
                'Azure Policy': AZURE_POLICIES_CACHE.get(control_id, []),  # Lista de tuplas
                'Security Principle': '',
                'Risk to mitigate': '',
                'MITRE ATT&CK': '',
                'Implementation example': '',
                'Criticality': '',
                'NIST SP 800-53 Rev.5': '',
                'PCI-DSS v4': '',
                'CIS Controls v8.1': '',
                'NIST CSF v2.0': '',
                'ISO 27001:2022': '',
                'SOC 2': ''
            }

            # Extraer secciones H3 del padre (y detectar controles hijo)
            child_controls = extract_parent_sections(h2, parent_data, domain_prefix)

            controls.append(parent_data)
            current_parent = parent_data

            # Agregar controles hijo que heredan Security Principle, Core Pillar y Azure Policy
            for child_data in child_controls:
                child_data['Security Principle'] = parent_data['Security Principle']
                child_data['Core Pillar'] = parent_data['Core Pillar']  # Heredar pillar
                child_data['Azure Policy'] = parent_data['Azure Policy']  # Heredar policy
                controls.append(child_data)
                print(f"    [HIJO] {child_data['Implementation ID']}: {child_data['Control Name']}")

        print(f"  Total controles extraídos: {len(controls)}")
        return controls

    except Exception as e:
        print(f"  ❌ ERROR: {str(e)}")
        return []

def extract_parent_sections(h2, data, domain_prefix):
    """Extrae secciones H3 de un control padre y detecta controles hijo"""
    current = h2.find_next_sibling()
    child_controls = []

    while current and current.name != 'h2':
        if current.name == 'h3':
            h3_text = current.get_text(strip=True)
            h3_title = h3_text.lower()

            # Detectar si H3 es un control hijo (formato "NS-1.1: Title")
            child_match = re.match(rf'^({domain_prefix}-\d+\.\d+)[:\s]+(.*)', h3_text, re.IGNORECASE)

            if child_match:
                # Es un control hijo
                child_full_id = child_match.group(1).upper()
                child_title = child_match.group(2).strip()

                # Usar ID completo (NS-1.1) en Implementation ID
                impl_id = child_full_id  # Completo: NS-1.1
                parent_id = '-'.join(child_full_id.split('.')[:1]) if '.' in child_full_id else child_full_id

                child_data = {
                    'Control ID': parent_id,  # Padre (NS-1)
                    'Implementation ID': impl_id,  # Hijo completo (NS-1.1)
                    'Control Name': child_title,
                    'Control Type': 'Child',
                    'Core Pillar': '',  # Se heredará del padre
                    'Azure Policy': [],  # Policies se heredarán del padre
                    'Security Principle': '',  # Se heredará del padre
                    'Risk to mitigate': '',
                    'MITRE ATT&CK': '',
                    'Implementation example': extract_h3_content(current)[:1500],
                    'Criticality': '',
                    'NIST SP 800-53 Rev.5': '',
                    'PCI-DSS v4': '',
                    'CIS Controls v8.1': '',
                    'NIST CSF v2.0': '',
                    'ISO 27001:2022': '',
                    'SOC 2': ''
                }
                child_controls.append(child_data)

            else:
                # Es una sección normal H3 del padre
                h3_content = extract_h3_content(current)

                if 'security principle' in h3_title:
                    data['Security Principle'] = h3_content[:1000]

                elif 'risk to mitigate' in h3_title:
                    formatted = h3_content[:1200]
                    formatted = re.sub(r'\. ([A-Z])', r'.\n\n\1', formatted)
                    data['Risk to mitigate'] = formatted

                elif 'mitre' in h3_title or 'att&ck' in h3_title:
                    mitre = h3_content[:1500]

                    # Formatear MITRE ATT&CK preservando palabras completas
                    # 1. Asegurar espacio después de punto antes de táctica
                    mitre = re.sub(r'\.([A-Z])', r'. \1', mitre)

                    # 2. Separar tácticas principales (TA####) con salto de línea ANTES
                    mitre = re.sub(r'([.!?])\s*([A-Z][a-z]+(?: [A-Z][a-z]+)? \(TA\d{4}\):)', r'\1\n\n\2', mitre)

                    # 3. Limpiar saltos excesivos y espacios
                    mitre = re.sub(r'\n{3,}', '\n\n', mitre)
                    mitre = re.sub(r' {2,}', ' ', mitre)

                    data['MITRE ATT&CK'] = mitre.strip()

                elif 'implementation example' in h3_title:
                    impl = h3_content[:2500]
                    # Saltos de línea en secciones clave
                    impl = re.sub(r'(Challenge:|Solution approach:|Solution:|Outcome:)', r'\n\n\1', impl)
                    impl = re.sub(r'\. ([A-Z][a-z]+:)', r'.\n\n\1', impl)
                    # Saltos después de punto + mayúscula
                    impl = re.sub(r'\.\s+([A-Z][a-z]+ [a-z])', r'.\n\n\1', impl)
                    data['Implementation example'] = impl.strip()

                elif 'criticality' in h3_title:
                    # Extraer el valor literal de la documentación
                    criticality_text = h3_content.strip()[:100].lower()
                    if 'must have' in criticality_text:
                        data['Criticality'] = 'Must have'
                    elif 'should have' in criticality_text:
                        data['Criticality'] = 'Should have'
                    elif 'nice to have' in criticality_text:
                        data['Criticality'] = 'Nice to have'
                    else:
                        # Usar texto literal si no coincide
                        data['Criticality'] = h3_content.strip()[:50]

                elif 'control mapping' in h3_title:
                    extract_framework_mappings(h3_content, data)

        current = current.find_next_sibling()

    return child_controls

def extract_h3_content(h3):
    """Extrae contenido después de un H3 hasta el siguiente H2/H3"""
    content = []
    current = h3.find_next_sibling()

    while current and current.name not in ['h2', 'h3']:
        text = current.get_text(strip=True)
        if text:
            content.append(text)
            if len(' '.join(content)) > 2000:
                break
        current = current.find_next_sibling()

    return ' '.join(content)

def extract_framework_mappings(h3_content, data):
    """Extrae control mappings de frameworks (con colon opcional)"""
    # IMPORTANTE: Colons son OPCIONALES - diferentes dominios usan diferentes formatos
    framework_patterns = {
        'CIS Controls v8.1': r'CIS Controls v[\d.]+\s*:?\s*([^N]+?)(?=(?:NIST|PCI|ISO|SOC\s+2|$))',
        'NIST SP 800-53 Rev.5': r'NIST SP 800-53 Rev\.?\s*\d*\s*:?\s*([^N]+?)(?=(?:PCI|CIS|ISO|SOC\s+2|$))',
        'PCI-DSS v4': r'PCI-DSS v[\d.]+\s*:?\s*([^N]+?)(?=(?:CIS|NIST|ISO|SOC\s+2|$))',
        'NIST CSF v2.0': r'(?:NIST Cybersecurity Framework|NIST CSF) v[\d.]+\s*:?\s*([^N]+?)(?=(?:ISO|SOC\s+2|$))',
        'ISO 27001:2022': r'ISO 27001:?\s*\d*\s*:?\s*([^N]+?)(?=(?:SOC\s+2|$))',
        'SOC 2': r'SOC 2\s*:?\s*(.+?)$'
    }

    for key, pattern in framework_patterns.items():
        match = re.search(pattern, h3_content, re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            value = re.sub(r'\s+', ' ', value).strip()
            value = value.rstrip(':').strip()
            # Validar que no contenga nombres de otros frameworks
            if value and len(value) < 200 and not any(fw in value for fw in ['PCI-DSS', 'CIS Controls', 'NIST Cybersecurity', 'NIST CSF', 'NIST SP', 'ISO 27001', 'SOC 2']):
                data[key] = value

def create_excel(all_controls, output_file):
    """Genera Excel con todos los controles"""
    print(f"\n{'='*80}")
    print("Generando Excel...")
    print(f"{'='*80}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Agrupar controles por dominio
    domains_controls = {}
    for control in all_controls:
        domain_prefix = control['Control ID'].split('-')[0]
        domain_name = next((d[0] for d in DOMAINS if d[2] == domain_prefix), domain_prefix)
        # Usar solo el nombre del dominio sin el prefijo
        sheet_name = domain_name

        if sheet_name not in domains_controls:
            domains_controls[sheet_name] = []
        domains_controls[sheet_name].append(control)

    # Crear hojas por dominio
    for sheet_name, controls in sorted(domains_controls.items()):
        ws = wb.create_sheet(title=sheet_name[:31])  # Límite 31 caracteres

        # Headers - NUEVO ORDEN con Implementation ID y Core Pillar
        headers = [
            'Control ID', 'Implementation ID', 'Control Name', 'Control Type', 'Core Pillar', 'Azure Policy',
            'Security Principle', 'Risk to mitigate', 'MITRE ATT&CK',
            'Implementation example', 'Criticality',
            'NIST SP 800-53 Rev.5', 'PCI-DSS v4', 'CIS Controls v8.1',
            'NIST CSF v2.0', 'ISO 27001:2022', 'SOC 2'
        ]

        ws.append(headers)

        # Formatear headers
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Datos
        for control in controls:
            # Preparar fila con valores
            row_values = []
            for h in headers:
                value = control.get(h, '')
                # Azure Policy es lista de tuplas, convertir temporalmente
                if h == 'Azure Policy' and isinstance(value, list):
                    row_values.append('TEMP_POLICIES')  # Marcador temporal
                else:
                    row_values.append(value)

            row_num = ws.max_row + 1
            ws.append(row_values)

            # Agregar hyperlink al Control ID
            control_id = control.get('Control ID', '')
            if control_id:
                control_id_col = headers.index('Control ID') + 1
                cell = ws.cell(row_num, control_id_col)

                # Extraer prefix del control (ej: AI-1 → AI)
                prefix = control_id.split('-')[0] if '-' in control_id else ''

                # Obtener domain slug del mapeo
                if prefix in DOMAIN_SLUG_MAP:
                    domain_slug = DOMAIN_SLUG_MAP[prefix]
                    mapping_url = f"https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-{domain_slug}#{control_id.lower()}"
                    cell.hyperlink = mapping_url
                    cell.font = Font(color="0563C1", underline="single")

            # Manejar Azure Policies
            azure_policy_col = headers.index('Azure Policy') + 1
            cell = ws.cell(row_num, azure_policy_col)

            if control.get('Azure Policy') and isinstance(control['Azure Policy'], list) and len(control['Azure Policy']) > 0:
                cell = ws.cell(row_num, azure_policy_col)
                policies = control['Azure Policy']
                control_id = control.get('Control ID', '')
                impl_id = control.get('Implementation ID', '')

                # Para child controls, usar el parent ID (ej: AI-1.1 → AI-1)
                if impl_id and '.' in impl_id:
                    # Es un child, extraer parent ID del Implementation ID
                    parent_id = impl_id.split('.')[0]
                else:
                    # Es parent, usar Control ID directamente
                    parent_id = control_id

                if len(policies) == 1:
                    # Una sola policy: nombre + hyperlink a documentación
                    policy_name, _ = policies[0]
                    cell.value = policy_name
                    if parent_id:
                        mapping_url = f"https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-controls-policy-mapping#{parent_id.lower()}"
                        cell.hyperlink = mapping_url
                        cell.font = Font(color="0563C1", underline="single")
                else:
                    # Múltiples policies: texto con hyperlink a documentación
                    policy_text = '\n'.join([name for name, _ in policies])
                    cell.value = policy_text

                    # Hyperlink a la documentación de control mapping (usar parent ID)
                    if parent_id:
                        mapping_url = f"https://learn.microsoft.com/en-us/security/benchmark/azure/mcsb-v2-controls-policy-mapping#{parent_id.lower()}"
                        cell.hyperlink = mapping_url
                        cell.font = Font(color="0563C1", underline="single")

                    # Agregar comentario con todas las URLs individuales del portal
                    comment_text = 'Azure Policy URLs:\n\n'
                    for policy_name, policy_url in policies:
                        if policy_url:
                            comment_text += f"• {policy_name}\n  {policy_url}\n\n"
                        else:
                            comment_text += f"• {policy_name}\n\n"

                    from openpyxl.comments import Comment
                    cell.comment = Comment(comment_text[:500], "Azure Policy Mapping")
            else:
                # No hay policies disponibles
                cell.value = "No Azure Policy available"
                cell.font = Font(color="666666", italic=True)

        # Ajustar anchos - ACTUALIZADO con Core Pillar
        # Control ID | Impl ID | Name | Type | Core Pillar | Azure Policy | Security | Risk | MITRE | Impl | Crit | NIST | PCI | CIS | NIST CSF | ISO | SOC
        column_widths = [12, 15, 45, 10, 25, 50, 50, 50, 50, 60, 10, 25, 20, 20, 25, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Altura de filas y wrap text
        ws.row_dimensions[1].height = 40
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 100
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        print(f"  ✓ {sheet_name}: {len(controls)} controles")

    # Readme
    readme = wb.create_sheet(title="Readme", index=0)
    readme['A1'] = "Microsoft Cloud Security Benchmark v2"
    readme['A1'].font = Font(bold=True, size=16)
    readme['A3'] = f"Generado: {time.strftime('%Y-%m-%d %H:%M:%S')}"
    readme['A4'] = "Fuente: https://learn.microsoft.com/en-us/security/benchmark/azure/overview"
    readme['A6'] = "Estructura:"
    readme['A7'] = "  - Controles Padre (ej: NS-1): Incluyen Azure Policy, Security Principle, Risk, MITRE, Criticality, Control mapping"
    readme['A8'] = "  - Controles Hijo (ej: 1.1, 1.2): Heredan Security Principle y Azure Policy del padre"
    readme['A9'] = "  - Columnas: Control ID (padre) + Implementation ID (hijo) para estructura clara"
    readme['A11'] = f"Total controles: {len(all_controls)}"
    readme['A12'] = f"Total dominios: {len(domains_controls)}"

    wb.save(output_file)
    print(f"\n✓ Excel guardado: {output_file}")
    print(f"  Total controles: {len(all_controls)}")
    print(f"  Total dominios: {len(domains_controls)}")

def main():
    print("\n" + "="*80)
    print(" "*15 + "MCSB v2 - Extracción Jerárquica Padre-Hijo")
    print("="*80)

    # Cargar Azure Policies primero
    load_azure_policies()

    all_controls = []

    for domain_name, domain_slug, domain_prefix in DOMAINS:
        controls = extract_controls_from_domain(domain_name, domain_slug, domain_prefix)
        all_controls.extend(controls)
        time.sleep(1)  # Rate limiting

    if all_controls:
        output_file = "docs/assets/tables/Microsoft_cloud_security_benchmark_v2.xlsx"
        create_excel(all_controls, output_file)
        print(f"\n{'='*80}")
        print("✓ COMPLETED!")
        print(f"{'='*80}\n")
    else:
        print("\n❌ No se extrajeron controles\n")

if __name__ == "__main__":
    main()
